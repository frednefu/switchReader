"""通用 Excel 导出工具 — 基于 openpyxl 生成 StreamingResponse。"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse


def export_to_excel(
    headers: list[str],
    rows: list[list],
    filename: str,
    sheet_title: str = "Sheet1",
    col_widths: list[int] | None = None,
) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    for row in rows:
        ws.append(row)

    # 列宽：有指定则用指定，否则自动计算
    for i, header in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[col_letter].width = col_widths[i - 1]
        else:
            # 简单估算：header 字符数 + 4
            ws.column_dimensions[col_letter].width = min(max(len(str(header)) + 4, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
