import io
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from math import ceil
import openpyxl

from app.database import get_db
from app.models.switch import Switch
from app.models.scan_log import ScanLog, ScanStatus, TriggerType
from app.schemas.switch import (
    SwitchCreate, SwitchUpdate, SwitchOut, SwitchTestRequest,
    SwitchTestResponse, SwitchImportRow, SwitchImportResult,
)
from app.schemas.scan import ScanLogOut, PaginatedResponse
from app.api.deps import get_current_user, require_admin
from app.services.scanner_service import trigger_scan, test_snmp_connection
from app.services.scheduler_service import refresh_job

router = APIRouter(prefix="/switches", tags=["交换机"])


@router.get("", response_model=PaginatedResponse)
def list_switches(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    search: str = Query("", max_length=128),
    is_active: bool = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = db.query(Switch)
    if search:
        q = q.filter(
            (Switch.name.contains(search)) | (Switch.ip_address.contains(search))
        )
    if is_active is not None:
        q = q.filter(Switch.is_active == is_active)
    total = q.count()
    pages = ceil(total / size) if total > 0 else 0
    items = q.order_by(Switch.id).offset((page - 1) * size).limit(size).all()
    return PaginatedResponse(
        items=[SwitchOut.model_validate(s).model_dump() for s in items],
        total=total, page=page, size=size, pages=pages,
    )


@router.post("", response_model=SwitchOut)
def create_switch(body: SwitchCreate, db: Session = Depends(get_db), admin=Depends(require_admin)):
    existing = db.query(Switch).filter(Switch.ip_address == body.ip_address).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="交换机 IP 已存在")
    sw = Switch(**body.model_dump(), created_by=admin.id)
    db.add(sw)
    db.commit()
    db.refresh(sw)
    if sw.is_active and sw.scan_interval > 0:
        refresh_job(sw.id, sw.scan_interval)
    return SwitchOut.model_validate(sw)


@router.get("/template")
def download_template(current_user=Depends(get_current_user)):
    """Download an Excel template for batch import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交换机导入模板"
    headers = ["名称", "IP地址", "SNMP团体字", "MIB类型", "SNMP端口", "扫描间隔(秒)"]
    ws.append(headers)
    # Example row
    ws.append(["示例-核心交换机", "192.168.1.1", "public", "huawei", 161, 86400])
    # Column widths
    widths = [20, 16, 16, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=switch_import_template.xlsx"},
    )


@router.post("/import", response_model=SwitchImportResult)
async def import_switches(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .csv 格式")

    content = await file.read()
    rows = []

    if file.filename.endswith('.csv'):
        import csv
        reader = csv.DictReader(io.StringIO(content.decode('utf-8-sig')))
        for r in reader:
            rows.append({
                "name": r.get("名称", "").strip(),
                "ip_address": r.get("IP地址", "").strip(),
                "community": r.get("SNMP团体字", "").strip(),
                "mib_type": r.get("MIB类型", "standard").strip() or "standard",
                "snmp_port": int(r.get("SNMP端口", 161) or 161),
                "scan_interval": int(r.get("扫描间隔(秒)", 86400) or 86400),
            })
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            rows.append({
                "name": str(row[0]).strip(),
                "ip_address": str(row[1]).strip(),
                "community": str(row[2]).strip() if row[2] else "",
                "mib_type": str(row[3] or "standard").strip() or "standard",
                "snmp_port": int(row[4] or 161) if row[4] else 161,
                "scan_interval": int(row[5] or 86400) if row[5] else 86400,
            })

    created = 0
    skipped = 0
    errors = []
    seen_ips = set()

    for r in rows:
        if not r["name"] or not r["ip_address"] or not r["community"]:
            errors.append(f"缺少必填字段: {r}")
            continue
        if r["ip_address"] in seen_ips:
            skipped += 1
            continue
        seen_ips.add(r["ip_address"])
        existing = db.query(Switch).filter(Switch.ip_address == r["ip_address"]).first()
        if existing:
            skipped += 1
            continue
        try:
            sw = Switch(**r, created_by=admin.id)
            db.add(sw)
            created += 1
        except Exception as e:
            errors.append(f"{r['ip_address']}: {e}")

    db.commit()
    return SwitchImportResult(created=created, skipped=skipped, errors=errors)


@router.get("/{switch_id}", response_model=SwitchOut)
def get_switch(switch_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    sw = db.query(Switch).get(switch_id)
    if not sw:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="交换机不存在")
    return SwitchOut.model_validate(sw)


@router.put("/{switch_id}", response_model=SwitchOut)
def update_switch(switch_id: int, body: SwitchUpdate, db: Session = Depends(get_db), admin=Depends(require_admin)):
    sw = db.query(Switch).get(switch_id)
    if not sw:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="交换机不存在")
    for key, val in body.model_dump(exclude_unset=True).items():
        setattr(sw, key, val)
    db.commit()
    db.refresh(sw)
    refresh_job(sw.id, sw.scan_interval if sw.is_active else 0)
    return SwitchOut.model_validate(sw)


@router.delete("/{switch_id}")
def delete_switch(switch_id: int, db: Session = Depends(get_db), admin=Depends(require_admin)):
    sw = db.query(Switch).get(switch_id)
    if not sw:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="交换机不存在")
    db.delete(sw)
    db.commit()
    refresh_job(switch_id, 0)
    return {"message": "交换机已删除"}


@router.post("/{switch_id}/scan", response_model=ScanLogOut)
async def trigger_switch_scan(
    switch_id: int,
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    sw = db.query(Switch).get(switch_id)
    if not sw:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="交换机不存在")

    running = db.query(ScanLog).filter(
        ScanLog.switch_id == switch_id,
        ScanLog.status == ScanStatus.running,
    ).first()
    if running:
        running.status = ScanStatus.failed
        running.error_message = "上次扫描意外中断，已自动重置"
        db.commit()

    scan_log_id = await trigger_scan(sw, TriggerType.manual)
    scan_log = db.query(ScanLog).get(scan_log_id)
    return ScanLogOut.model_validate(scan_log)


# ─── 取消扫描 ───

@router.post("/{switch_id}/cancel-scan")
def cancel_scan(
    switch_id: int,
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    sw = db.query(Switch).get(switch_id)
    if not sw:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="交换机不存在")
    running = db.query(ScanLog).filter(
        ScanLog.switch_id == switch_id,
        ScanLog.status == ScanStatus.running,
    ).first()
    if not running:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前没有正在进行的扫描")
    running.status = ScanStatus.failed
    running.error_message = "用户手动取消"
    db.commit()
    return {"message": "扫描已取消"}


@router.post("/scan-all")
async def scan_all_switches(
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    """扫描所有启用的交换机。"""
    switches = db.query(Switch).filter(Switch.is_active == True).all()
    if not switches:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="没有可扫描的交换机")
    started = 0
    for sw in switches:
        running = db.query(ScanLog).filter(
            ScanLog.switch_id == sw.id,
            ScanLog.status == ScanStatus.running,
        ).first()
        if running:
            running.status = ScanStatus.failed
            running.error_message = "上次扫描意外中断，已自动重置"
            db.commit()
        await trigger_scan(sw, TriggerType.manual)
        started += 1
    return {"message": f"已触发 {started} 台交换机扫描",
            "started": started}


@router.delete("/all")
def delete_all_switches(
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    """删除所有交换机及其关联数据。"""
    count = db.query(Switch).count()
    if count == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="没有可删除的交换机")
    db.query(Switch).delete()
    db.commit()
    return {"message": f"已删除 {count} 台交换机及关联数据", "deleted": count}


@router.post("/test", response_model=SwitchTestResponse)
async def test_switch_connection(
    body: SwitchTestRequest,
    current_user=Depends(get_current_user),
):
    result = await test_snmp_connection(body.ip_address, body.community)
    return result
